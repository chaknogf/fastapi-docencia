from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session as SQLAlchemySession
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import asc, desc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import tempfile
from pathlib import Path
import time

from app.apicore import get_db
from app.models.actividades import VistaActividad
from app.schemas.actividad import ActividadVista

router = APIRouter(tags=["reportes"])


def limpiar_reportes_antiguos(directorio: Path, horas: int = 24):
    try:
        ahora = time.time()
        limite = ahora - (horas * 3600)
        for archivo in directorio.glob("reporte_actividades_*.xlsx"):
            if archivo.stat().st_mtime < limite:
                archivo.unlink()
    except Exception:
        pass


@router.get("/reporte/excel")
async def generar_reporte_excel(
    anio: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    subdireccion_id: Optional[int] = Query(None),
    servicio_id: Optional[int] = Query(None),
    db: SQLAlchemySession = Depends(get_db),
):
    try:
        query = db.query(VistaActividad)

        if anio is not None:
            query = query.filter(VistaActividad.anio == anio)
        if mes is not None:
            query = query.filter(VistaActividad.mes_id == mes)
        if subdireccion_id is not None:
            query = query.filter(VistaActividad.subdireccion_id == subdireccion_id)
        if servicio_id is not None:
            query = query.filter(VistaActividad.servicio_id == servicio_id)

        actividades = query.order_by(
            asc(VistaActividad.fecha_programada),
            desc(VistaActividad.id),
        ).all()

        if not actividades:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron actividades con los filtros especificados",
            )

        temp_dir = Path(tempfile.gettempdir()) / "reportes_excel"
        temp_dir.mkdir(exist_ok=True)

        limpiar_reportes_antiguos(temp_dir, horas=24)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Actividades"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "ID",
            "Tema",
            "Actividad",
            "Servicio",
            "Responsable",
            "Fecha Programada",
            "Modalidad",
            "Estado",
            "Mes",
            "Año",
            "Fecha Entrega Informe",
            "Observaciones",
        ]

        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for act in actividades:
            persona_responsable = act.persona_responsable or {}
            responsable = persona_responsable.get("r0", {}).get("nombre", "")

            detalles = act.detalles or {}
            fecha_entrega = detalles.get("fecha_entrega_informe", "")
            observaciones = detalles.get("observaciones", "")

            row_data = [
                act.id,
                act.tema,
                act.actividad,
                act.servicio_encargado,
                responsable,
                act.fecha_programada.strftime("%Y-%m-%d")
                if act.fecha_programada
                else "",
                act.modalidad,
                act.estado,
                act.mes_id,
                act.anio,
                fecha_entrega,
                observaciones,
            ]

            ws.append(row_data)

        column_widths = {
            "A": 8,
            "B": 30,
            "C": 40,
            "D": 15,
            "E": 25,
            "F": 18,
            "G": 15,
            "H": 15,
            "I": 8,
            "J": 8,
            "K": 20,
            "L": 30,
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=1,
            max_col=len(headers),
        ):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_actividades_{timestamp}.xlsx"
        filepath = temp_dir / filename

        wb.save(str(filepath))

        return FileResponse(
            path=str(filepath),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except HTTPException:
        raise
    except SQLAlchemyError as e:
        raise HTTPException(status_code=500, detail=f"Error de base de datos: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte: {str(e)}")
